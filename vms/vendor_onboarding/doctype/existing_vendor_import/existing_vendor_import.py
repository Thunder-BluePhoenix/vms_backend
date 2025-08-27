# Enhanced existing_vendor_import.py
# Copyright (c) 2025, Blue Phoenix and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
import pandas as pd
import json
import os
import io
from frappe.utils import cstr, flt, cint, today, now, get_site_path, validate_email_address
from frappe.utils.file_manager import get_file
import zipfile
from io import BytesIO
import re
import openpyxl
from frappe.utils.file_manager import save_file

try:
    from .existing_vendor_import_utils import VendorImportUtils
except ImportError:
    # Alternative import path if the above fails
    from vms.vendor_onboarding.doctype.existing_vendor_import.existing_vendor_import_utils import VendorImportUtils

# Also add this helper method to handle the utils class instantiation
def get_vendor_utils():
    """Get instance of VendorImportUtils class"""
    try:
        return VendorImportUtils()
    except Exception as e:
        frappe.log_error(f"Error instantiating VendorImportUtils: {str(e)}")
        return None
	


class ExistingVendorImport(Document):
	def validate(self):
		if self.csv_xl and not self.existing_vendor_initialized:
			self.parse_and_validate_data()

	def parse_and_validate_data(self):
		"""Parse CSV/Excel file and validate data"""
		try:
			# Get file content
			file_doc = frappe.get_doc("File", {"file_url": self.csv_xl})
			file_path = file_doc.get_full_path()
			
			# Read file based on extension
			if file_path.endswith('.csv'):
				df = pd.read_csv(file_path)
			else:
				df = pd.read_excel(file_path)
			
			# Clean column names
			df.columns = df.columns.str.strip()
			
			# Store original headers for mapping
			self.original_headers = json.dumps(list(df.columns))
			
			# Convert DataFrame to list of dictionaries
			vendor_data = df.to_dict('records')
			
			# Store parsed data
			self.vendor_data = json.dumps(vendor_data, indent=2, default=str)
			
			# Generate field mapping interface
			self.generate_field_mapping_html(df.columns)
			
			# Auto-apply mapping if not manually set
			if not self.field_mapping:
				auto_mapping = self.generate_auto_mapping(df.columns)
				self.field_mapping = json.dumps(auto_mapping, indent=2)
			
			# Generate mapping statistics
			self.mapping_statistics = self.generate_mapping_statistics_html()
			
			# Validate and process data
			validation_results = self.validate_vendor_data(vendor_data)
			self.success_fail_rate = json.dumps(validation_results, indent=2)
			
			# Generate HTML for display
			self.generate_display_html(vendor_data, validation_results)
			
		except Exception as e:
			frappe.throw(f"Error parsing file: {str(e)}")

	def generate_auto_mapping(self, csv_headers):
		"""Generate automatic field mapping based on header similarity"""
		
		# Define comprehensive mapping patterns
		mapping_patterns = {
			# Vendor Master fields - more comprehensive patterns
			'vendor_name': [
				'vendor name', 'company name', 'supplier name', 'name', 'vendor_name',
				'vendorname', 'company_name', 'supplier_name', 'firm name', 'organization name'
			],
			'office_email_primary': [
				'primary email', 'email-id', 'email', 'contact email', 'office email',
				'email_id', 'emailid', 'email address', 'primary_email', 'main email'
			],
			'office_email_secondary': [
				'secondary email', 'alternate email', 'backup email', 'secondary_email',
				'alt email', 'email 2', 'second email'
			],
			'mobile_number': [
				'contact no', 'phone', 'mobile', 'telephone', 'contact number',
				'contact_no', 'mobile_number', 'phone_number', 'contact_number',
				'mobile no', 'cell phone', 'phone no'
			],
			'country': ['country'],
			
			# Company Vendor Code fields
			'company_code': [
				'c.code', 'company code', 'comp code', 'company_code',
				'ccode', 'c code', 'company id', 'comp_code'
			],
			'vendor_code': [
				'vendor code', 'supplier code', 'sap code', 'vendor_code',
				'vendorcode', 'supplier_code', 'sap_code', 'vendor id'
			],
			'gst_no': [
				'gstn no', 'gst no', 'gst number', 'gstin', 'gst_no',
				'gstn_no', 'gst_number', 'tax_number', 'gstin_no'
			],
			'state': ['state', 'state name', 'state_name'],
			'state_code': ['state code', 'state_code', 'statecode'],
			
			# Company Details fields
			'company_pan_number': [
				'pan no', 'pan number', 'pan', 'company_pan_number',
				'pan_no', 'pan_number', 'company pan', 'tax_id'
			],
			'address_line_1': [
				'address01', 'address 1', 'address line 1', 'street 1',
				'address1', 'address_line_1', 'street address', 'address'
			],
			'address_line_2': [
				'address02', 'address 2', 'address line 2', 'street 2',
				'address2', 'address_line_2', 'address line 2'
			],
			'city': ['city', 'city_name'],
			'pincode': [
				'pincode', 'pin code', 'postal code', 'zip', 'pin_code',
				'zipcode', 'zip_code', 'postal_code'
			],
			'telephone_number': [
				'telephone number', 'landline', 'office phone', 'telephone_number',
				'landline_number', 'office_phone', 'tel_no'
			],
			
			# Multiple Company Data fields
			'purchase_organization': [
				'purchase organization', 'purchase org', 'po', 'porg',
				'purchase_organization', 'purchase_org', 'purch_org'
			],
			'account_group': [
				'account group', 'acc group', 'account grp', 'account_group',
				'acc_group', 'account_grp'
			],
			'terms_of_payment': [
				'terms of payment', 'payment terms', 'payment condition',
				'terms_of_payment', 'payment_terms', 'payment_condition'
			],
			'purchase_group': [
				'purchase group', 'purchasing group', 'purch group',
				'purchase_group', 'purchasing_group', 'purch_group'
			],
			'order_currency': [
				'order currency', 'currency', 'curr', 'order_currency',
				'payment_currency', 'transaction_currency'
			],
			'incoterms': [
				'incoterm', 'incoterms', 'delivery terms', 'inco_terms',
				'delivery_terms', 'shipping_terms'
			],
			'reconciliation_account': [
				'reconciliation account', 'recon account', 'gl account',
				'reconciliation_account', 'recon_account', 'gl_account'
			],
			
			# Payment Details fields
			'bank_name': ['bank name', 'bank_name', 'bank'],
			'ifsc_code': ['ifsc code', 'ifsc', 'ifsc_code', 'bank_code'],
			'account_number': [
				'account number', 'bank account', 'account_number',
				'bank_account', 'acc_number', 'account no'
			],
			'name_of_account_holder': [
				'name of account holder', 'account holder name',
				'name_of_account_holder', 'account_holder_name', 'account holder'
			],
			'type_of_account': [
				'type of account', 'account type', 'type_of_account',
				'account_type', 'acc_type'
			],
			
			# Additional fields
			'vendor_gst_classification': [
				'vendor gst classification', 'gst classification',
				'vendor_gst_classification', 'gst_classification'
			],
			'nature_of_services': [
				'nature of services', 'service type', 'nature_of_services',
				'service_type', 'services'
			],
			'vendor_type': [
				'vendor type', 'supplier type', 'vendor_type',
				'supplier_type', 'type'
			],
			'remarks': ['remarks', 'comments', 'notes', 'remark'],
			'established_year': [
				'established year', 'year established', 'established_year',
				'year_established', 'incorporation year'
			],
			'nature_of_business': [
				'nature of business', 'business nature', 'nature_of_business',
				'business_nature', 'business type'
			],
			'type_of_business': [
				'type of business', 'business type', 'type_of_business',
				'business_type'
			],
			'corporate_identification_number': [
				'cin', 'cin number', 'corporate identification',
				'corporate_identification_number', 'cin_number'
			]
		}
		
		auto_mapping = {}
		
		for csv_header in csv_headers:
			csv_header_lower = csv_header.lower().strip()
			mapped = False
			
			# Try to find exact or partial matches
			for field_name, patterns in mapping_patterns.items():
				for pattern in patterns:
					if (pattern == csv_header_lower or 
						pattern in csv_header_lower or 
						csv_header_lower in pattern):
						auto_mapping[csv_header] = field_name
						mapped = True
						break
				if mapped:
					break
			
			# If no mapping found, set to None
			if not mapped:
				auto_mapping[csv_header] = None
		
		return auto_mapping

	def process_vendors(self):
		"""Enhanced process vendors with standalone payment details tracking"""
		
		if not self.csv_xl:
			frappe.throw("Please upload a CSV/Excel file first")
		
		if not self.vendor_data:
			frappe.throw("No vendor data found. Please save the form to parse the CSV file first")
		
		if not self.field_mapping:
			frappe.throw("Please configure field mapping first")
		
		vendor_data = json.loads(self.vendor_data)
		field_mapping = json.loads(self.field_mapping)
		
		results = {
			"total_processed": 0,
			"vendors_created": 0,
			"vendors_updated": 0,
			"company_codes_created": 0,
			"company_codes_updated": 0,
			"payment_details_created": 0,      # Track standalone payment details
			"payment_details_updated": 0,      # Track updates
			"errors": [],
			"warnings": []
		}
		
		for idx, row in enumerate(vendor_data, 1):
			try:
				# Apply field mapping (preserves original row)
				mapped_row = self.apply_field_mapping(row, field_mapping)
				
				# Process vendor with standalone payment details
				vendor_result = self.process_single_vendor(mapped_row, idx)
				
				# Aggregate results
				results["total_processed"] += 1
				
				if vendor_result.get("vendor_action") == "created":
					results["vendors_created"] += 1
				elif vendor_result.get("vendor_action") == "updated":
					results["vendors_updated"] += 1
				
				if vendor_result.get("company_code_action") == "created":
					results["company_codes_created"] += 1
				elif vendor_result.get("company_code_action") == "updated":
					results["company_codes_updated"] += 1
				
				# Track payment details creation/updates
				if vendor_result.get("payment_details_action") == "created":
					results["payment_details_created"] += 1
				elif vendor_result.get("payment_details_action") == "updated":
					results["payment_details_updated"] += 1
				
				if vendor_result.get("warnings"):
					results["warnings"].extend(vendor_result["warnings"])
					
			except Exception as e:
				error_msg = f"Row {idx}: {str(e)}"
				results["errors"].append(error_msg)
				frappe.log_error(f"Vendor import error: {error_msg}")
		
		# Mark as completed
		self.existing_vendor_initialized = 1
		self.save()
		
		return results





	# FIXED PAYMENT DETAILS CREATION FOR VENDOR BANK DETAILS
# Replace your create_standalone_payment_details method in existing_vendor_import.py

	def create_standalone_payment_details(self, mapped_row, vendor_master_name):
		"""Create payment details directly using Vendor Bank Details doctype"""
		
		result = {
			'action': 'none',
			'warnings': [],
			'payment_doc_name': None
		}
		
		try:
			# Check if any payment-related fields are present
			payment_fields = [
				'bank_name', 'ifsc_code', 'account_number', 'name_of_account_holder', 'type_of_account',
				'beneficiary_name', 'beneficiary_swift_code', 'beneficiary_iban_no'
			]
			
			# Check mapped row for payment data
			has_payment_data = any(mapped_row.get(field) for field in payment_fields)
			
			# Also check original row for payment data
			if not has_payment_data:
				original_row = mapped_row.get('_original_row', {})
				payment_csv_fields = [
					'Bank Name', 'IFSC Code', 'Account Number', 'Name of Account Holder', 'Type of Account',
					'Beneficiary Name', 'Beneficiary Swift Code', 'Beneficiary IBAN No.'
				]
				has_payment_data = any(original_row.get(field) for field in payment_csv_fields)
			
			if not has_payment_data:
				result['warnings'].append(f"No payment data found for vendor {mapped_row.get('vendor_name')}")
				return result
			
			vendor_name = mapped_row.get('vendor_name', '').strip()
			company_code = mapped_row.get('company_code', '').strip()
			
			if not vendor_name:
				result['warnings'].append("Vendor name is required for payment details")
				return result
			
			# Check if payment details already exist for this vendor
			existing_payment = frappe.db.exists("Vendor Bank Details", {
				"ref_no": vendor_master_name
			})
			
			if existing_payment:
				payment_doc = frappe.get_doc("Vendor Bank Details", existing_payment)
				result['action'] = 'updated'
			else:
				payment_doc = frappe.new_doc("Vendor Bank Details")
				payment_doc.ref_no = vendor_master_name  # Link to Vendor Master
				result['action'] = 'created'
			
			# Import utilities for field mapping
			from .existing_vendor_import_utils import VendorImportUtils
			
			# Map basic payment fields
			payment_row = mapped_row.get('_original_row', mapped_row)
			
			# Track populated fields for logging
			populated_fields = []
			
			# Map direct payment fields (main form fields)
			basic_fields = {
				'bank_name': 'bank_name',
				'ifsc_code': 'ifsc_code',
				'account_number': 'account_number',
				'name_of_account_holder': 'name_of_account_holder',
				'type_of_account': 'type_of_account',
				'currency': 'currency'
			}
			
			for csv_field, doc_field in basic_fields.items():
				value = VendorImportUtils.get_flexible_field_value(payment_row, csv_field)
				if value:
					setattr(payment_doc, doc_field, value)
					populated_fields.append(f"{doc_field}: {value}")
			
			# Handle banker details (for domestic banking - use banker_details table)
			banker_data = VendorImportUtils.extract_banker_details(payment_row)
			if banker_data and any(banker_data.values()):
				# Clear existing banker details
				payment_doc.set('banker_details', [])
				# Add new banker detail
				payment_doc.append("banker_details", banker_data)
				populated_fields.append(f"Banker Details")
			
			# Handle international bank details
			intl_data = VendorImportUtils.extract_international_bank_details_for_vendor_bank(payment_row)
			if intl_data and any(intl_data.values()):
				# Clear existing international bank details
				payment_doc.set('international_bank_details', [])
				# Add new international bank detail
				payment_doc.append("international_bank_details", intl_data)
				populated_fields.append(f"International Bank Details")
			
			# Handle intermediate bank details (if needed)
			intermediate_data = VendorImportUtils.extract_intermediate_bank_details(payment_row)
			if intermediate_data and any(intermediate_data.values()):
				payment_doc.add_intermediate_bank_details = 1  # Enable intermediate bank details
				payment_doc.set('intermediate_bank_details', [])
				payment_doc.append("intermediate_bank_details", intermediate_data)
				populated_fields.append(f"Intermediate Bank Details")
			
			# Set transaction preferences if available
			if VendorImportUtils.get_flexible_field_value(payment_row, 'rtgs'):
				payment_doc.rtgs = 1
			if VendorImportUtils.get_flexible_field_value(payment_row, 'neft'):
				payment_doc.neft = 1
			if VendorImportUtils.get_flexible_field_value(payment_row, 'ift'):
				payment_doc.ift = 1
			
			# Set metadata
			if hasattr(payment_doc, 'created_from_import'):
				payment_doc.created_from_import = 1
			if hasattr(payment_doc, 'import_date'):
				payment_doc.import_date = frappe.utils.now()
			if hasattr(payment_doc, 'import_source'):
				payment_doc.import_source = "Existing Vendor Import"
			
			# Save the document
			if populated_fields:
				payment_doc.save(ignore_permissions=True)
				frappe.db.commit()
			

				frappe.db.set_value("Vendor Master", vendor_master_name, "bank_details", payment_doc.name)
				result['payment_doc_name'] = payment_doc.name
				# frappe.log_error(f"Payment details {result['action']}: {payment_doc.name} for {vendor_name}, Fields: {populated_fields}", "Payment Success")
			else:
				result['warnings'].append(f"No valid payment data found for {vendor_name}")
		
		except Exception as e:
			error_msg = f"Error creating payment details for {mapped_row.get('vendor_name', 'Unknown')}: {str(e)}"
			frappe.log_error(error_msg, "Payment Details Creation Error")
			result['warnings'].append(error_msg)
		
		return result

	def generate_payment_ref_no(self, vendor_name, company_code=""):
		"""Generate unique reference number for standalone payment details"""
		
		try:
			# Create a simple reference based on vendor name and company
			vendor_short = ''.join([c for c in vendor_name.upper() if c.isalnum()])[:8]
			company_short = company_code[:4] if company_code else "COMP"
			
			# Get current timestamp
			now = frappe.utils.now_datetime()
			timestamp = now.strftime("%y%m%d%H%M")
			
			# Create pattern: PAY-{vendor_short}-{company_short}-{timestamp}
			base_ref = f"PAY-{vendor_short}-{company_short}-{timestamp}"
			
			# Ensure uniqueness
			ref_no = base_ref
			counter = 1
			while frappe.db.exists("Vendor Bank Details", {"ref_no": ref_no}):
				ref_no = f"{base_ref}-{counter:02d}"
				counter += 1
			
			return ref_no
			
		except Exception as e:
			frappe.log_error(f"Error generating payment ref_no: {str(e)}")
			# Fallback to timestamp-based ref_no
			timestamp = frappe.utils.now_datetime().strftime("%Y%m%d%H%M%S")
			return f"PAY-IMP-{timestamp}"









	def process_single_vendor(self, mapped_row, row_number):
		"""Enhanced process_single_vendor with direct payment details creation"""
		
		vendor_name = str(mapped_row.get('vendor_name', '')).strip()
		vendor_code = str(mapped_row.get('vendor_code', '')).strip()
		company_code = str(mapped_row.get('company_code', '')).strip()
		state = str(mapped_row.get('state', '')).strip()
		gst_no = str(mapped_row.get('gst_no', '')).strip()
		
		if not vendor_name:
			raise frappe.ValidationError(f"Vendor name is required")
		
		result = {
			"vendor_action": None,
			"company_code_action": None,
			"payment_details_action": None,
			"warnings": []
		}
		
		# Step 1: Find or create Vendor Master
		vendor_master = self.find_or_create_vendor_master(mapped_row)
		result["vendor_action"] = "updated" if frappe.db.exists("Vendor Master", {"vendor_name": vendor_name}) else "created"
		
		# Step 2: Handle Company Vendor Code with enhanced duplicate logic
		if vendor_code and company_code:
			company_code_result = self.handle_company_vendor_code(
				vendor_master.name, 
				mapped_row, 
				vendor_code, 
				company_code, 
				state, 
				gst_no
			)
			result.update(company_code_result)
		
		# Step 3: Create/update company details
		self.create_vendor_company_details(vendor_master.name, mapped_row)
		
		# Step 4: Create/update multiple company data
		self.create_multiple_company_data(vendor_master.name, mapped_row)

		payment_result = self.create_standalone_payment_details(mapped_row, vendor_master.name)
		result["payment_details_action"] = payment_result.get('action', 'none')
			
		if payment_result.get('warnings'):
			result["warnings"].extend(payment_result['warnings'])
		
		return result
	

	def process_single_vendor_dup(self, mapped_row, row_number):
		"""Enhanced process_single_vendor with direct payment details creation"""
		
		vendor_name = str(mapped_row.get('vendor_name', '')).strip()
		vendor_code = str(mapped_row.get('vendor_code', '')).strip()
		company_code = str(mapped_row.get('company_code', '')).strip()
		state = str(mapped_row.get('state', '')).strip()
		gst_no = str(mapped_row.get('gst_no', '')).strip()
		
		if not vendor_name:
			raise frappe.ValidationError(f"Vendor name is required")
		
		result = {
			"vendor_action": None,
			"company_code_action": None,
			"payment_details_action": None,
			"warnings": []
		}
		
		try:
			# Step 1: Find or create Vendor Master
			vendor_master = self.find_or_create_vendor_master(mapped_row)
			vendor_exists = frappe.db.exists("Vendor Master", {"vendor_name": vendor_name})
			result["vendor_action"] = "updated" if vendor_exists else "created"
			
			# Step 2: Handle Company Vendor Code
			company_vendor_code = self.find_or_create_vendor_master(mapped_row, vendor_master.name)
			result["company_code_action"] = "updated" if company_vendor_code.get('exists') else "created"
			
			# Step 3: Create or update Company Details  
			company_details = self.create_vendor_company_details( vendor_master.name, mapped_row)
			
			# Step 4: Create or update Multiple Company Data
			self.create_multiple_company_data(vendor_master.name, mapped_row )
			
			# Step 5: CREATE PAYMENT DETAILS DIRECTLY (No Vendor Onboarding needed)
			payment_result = self.create_standalone_payment_details(mapped_row, vendor_master.name)
			result["payment_details_action"] = payment_result.get('action', 'none')
			
			if payment_result.get('warnings'):
				result["warnings"].extend(payment_result['warnings'])
			
			frappe.db.commit()
			
		except Exception as e:
			frappe.db.rollback()
			frappe.log_error(f"Error processing vendor {vendor_name}: {str(e)}")
			raise
		
		return result

	def find_or_create_vendor_master(self, mapped_row):
		"""Find existing vendor or create new one"""
		
		vendor_name = str(mapped_row.get('vendor_name', '')).strip()
		office_email = str(mapped_row.get('office_email_primary', '')).strip()
		
		# Try to find existing vendor by name or email
		existing_vendor = None
		
		# Search by vendor name first
		if vendor_name:
			existing_vendor = frappe.db.exists("Vendor Master", {"vendor_name": vendor_name})
		
		# If not found by name, search by email
		if not existing_vendor and office_email:
			existing_vendor = frappe.db.exists("Vendor Master", {"office_email_primary": office_email})
		
		if existing_vendor:
			# Update existing vendor
			vendor_master = frappe.get_doc("Vendor Master", existing_vendor)
			self.update_vendor_master_fields(vendor_master, mapped_row)
		else:
			# Create new vendor
			vendor_master = frappe.new_doc("Vendor Master")
			self.set_vendor_master_fields(vendor_master, mapped_row)
		
		vendor_master.save(ignore_permissions=True)
		return vendor_master

	def handle_company_vendor_code(self, vendor_ref_no, mapped_row, vendor_code, company_code, state, gst_no):
		"""Enhanced handling of Company Vendor Code with proper duplicate logic"""
		
		result = {
			"company_code_action": None,
			"warnings": []
		}
		
		# Find company master
		company_master = frappe.db.exists("Company Master", {"company_code": company_code})
		if not company_master:
			result["warnings"].append(f"Company with code {company_code} not found. Please create company master first.")
			return result
		
		company_doc = frappe.get_doc("Company Master", company_master)
		
		# Check if Company Vendor Code exists for this vendor + company combination
		existing_cvc = frappe.db.exists("Company Vendor Code", {
			"vendor_ref_no": vendor_ref_no,
			"company_name": company_doc.name
		})
		
		if existing_cvc:
			# Update existing Company Vendor Code
			cvc_doc = frappe.get_doc("Company Vendor Code", existing_cvc)
			
			# Check if this vendor code + state + GST combination already exists
			duplicate_found = False
			
			for vc_row in cvc_doc.vendor_code:
				if (str(vc_row.vendor_code).strip() == vendor_code and 
					str(vc_row.state).strip() == state and 
					str(vc_row.gst_no).strip() == gst_no):
					duplicate_found = True
					result["warnings"].append(f"Vendor code {vendor_code} for state {state} with GST {gst_no} already exists")
					break
			
			# If no duplicate, add new vendor code row
			if not duplicate_found:
				cvc_doc.append("vendor_code", {
					"vendor_code": vendor_code,
					"state": state,
					"gst_no": gst_no
				})
				result["company_code_action"] = "updated"
			
		else:
			# Create new Company Vendor Code
			cvc_doc = frappe.new_doc("Company Vendor Code")
			cvc_doc.vendor_ref_no = vendor_ref_no
			cvc_doc.company_name = company_doc.name
			
			# Add vendor code row
			cvc_doc.append("vendor_code", {
				"vendor_code": vendor_code,
				"state": state,
				"gst_no": gst_no
			})
			result["company_code_action"] = "created"
		
		cvc_doc.save(ignore_permissions=True)
		return result

	def create_vendor_company_details(self, vendor_ref_no, mapped_row):
		"""Create or update vendor company details"""
		
		company_code = str(mapped_row.get('company_code', '')).strip()
		if not company_code:
			return
		
		# Find company master
		company_master = frappe.db.exists("Company Master", {"company_code": company_code})
		if not company_master:
			return
		
		# Check if company details already exist
		existing_details = frappe.db.exists("Vendor Onboarding Company Details", {
			"ref_no": vendor_ref_no,
			"company_name": company_master
		})
		
		if not existing_details:
			# Create new company details
			company_details = frappe.new_doc("Vendor Onboarding Company Details")
			company_details.ref_no = vendor_ref_no
			company_details.company_name = company_master
			
			# Set company detail fields with proper string conversion
			field_mapping = {
				'gst': 'gst',
				'company_pan_number': 'company_pan_number',
				'address_line_1': 'address_line_1',
				'address_line_2': 'address_line_2',
				'city': 'city',
				'state': 'state',
				'country': 'country',
				'pincode': 'pincode',
				'telephone_number': 'telephone_number',
				'nature_of_business': 'nature_of_business',
				'type_of_business': 'type_of_business',
				'corporate_identification_number': 'corporate_identification_number',
				'established_year': 'established_year'
			}
			
			for csv_field, doc_field in field_mapping.items():
				if mapped_row.get(csv_field):
					# Convert to string and clean
					value = str(mapped_row[csv_field]).strip() if mapped_row[csv_field] not in [None, 'nan', 'NaN'] else None
					if value and value.lower() not in ['nan', 'none', 'null']:
						company_details.set(doc_field, value)
			
			# Create or link master data
			if mapped_row.get('state'):
				state_value = str(mapped_row['state']).strip()
				if state_value and state_value.lower() not in ['nan', 'none', 'null']:
					company_details.state = self.get_or_create_state(state_value)
			
			if mapped_row.get('city'):
				city_value = str(mapped_row['city']).strip()
				if city_value and city_value.lower() not in ['nan', 'none', 'null']:
					company_details.city = self.get_or_create_city(city_value)
			
			if mapped_row.get('pincode'):
				pincode_value = str(mapped_row['pincode']).strip()
				if pincode_value and pincode_value.lower() not in ['nan', 'none', 'null']:
					company_details.pincode = self.get_or_create_pincode(pincode_value)
			
			company_details.save(ignore_permissions=True)

	def create_multiple_company_data(self, vendor_ref_no, mapped_row):
		"""Create or update multiple company data"""
		
		company_code = str(mapped_row.get('company_code', '')).strip()
		if not company_code:
			return
		
		# Find company master
		company_master = frappe.db.exists("Company Master", {"company_code": company_code})
		if not company_master:
			return
		
		# Get vendor master
		vendor_master = frappe.get_doc("Vendor Master", vendor_ref_no)
		
		# Check if multiple company data already exists
		duplicate_found = False
		for mcd in vendor_master.multiple_company_data:
			if mcd.company_name == company_master:
				duplicate_found = True
				break
		
		if not duplicate_found:
			# Add new multiple company data row
			mcd_row = {
				"company_name": company_master,
				"purchase_organization": self.safe_get_value(mapped_row, 'purchase_organization'),
				"account_group": self.safe_get_value(mapped_row, 'account_group'),
				"terms_of_payment": self.safe_get_value(mapped_row, 'terms_of_payment'),
				"purchase_group": self.safe_get_value(mapped_row, 'purchase_group'),
				"order_currency": self.safe_get_value(mapped_row, 'order_currency'),
				"incoterms": self.safe_get_value(mapped_row, 'incoterms'),
				"reconciliation_account": self.safe_get_value(mapped_row, 'reconciliation_account')
			}

			
			
			vendor_master.append("multiple_company_data", mcd_row)
			
			vendor_master.save(ignore_permissions=True)

	def safe_get_value(self, mapped_row, field_name):
		"""Safely get and convert value to string"""
		value = mapped_row.get(field_name)
		if value is None or str(value).lower() in ['nan', 'none', 'null', '']:
			return None
		return str(value).strip()

	def set_vendor_master_fields(self, vendor_master, mapped_row):
		"""Set fields for new vendor master"""
		
		field_mapping = {
			'vendor_name': 'vendor_name',
			'office_email_primary': 'office_email_primary',
			'office_email_secondary': 'office_email_secondary',
			'mobile_number': 'mobile_number',
			'country': 'country',
			'payee_in_document': 'payee_in_document',
			'gr_based_inv_ver': 'gr_based_inv_ver',
			'service_based_inv_ver': 'service_based_inv_ver',
			'check_double_invoice': 'check_double_invoice'
		}
		
		for csv_field, doc_field in field_mapping.items():
			value = self.safe_get_value(mapped_row, csv_field)
			if value:
				vendor_master.set(doc_field, value)

		vtype_row = {
				"vendor_type": self.safe_get_value(mapped_row, 'vendor_type')
			}
		vendor_master.append("vendor_types", vtype_row)
		# Set default values for checkboxes if not provided
		vendor_master.payee_in_document = 1 if not mapped_row.get('payee_in_document') else self.safe_get_value(mapped_row, 'payee_in_document')
		vendor_master.gr_based_inv_ver = 1 if not mapped_row.get('gr_based_inv_ver') else self.safe_get_value(mapped_row, 'gr_based_inv_ver')
		vendor_master.service_based_inv_ver = 1 if not mapped_row.get('service_based_inv_ver') else self.safe_get_value(mapped_row, 'service_based_inv_ver')
		vendor_master.check_double_invoice = 1 if not mapped_row.get('check_double_invoice') else self.safe_get_value(mapped_row, 'check_double_invoice')
		
		# Set additional fields
		vendor_master.status = "Active"
		vendor_master.registered_date = today()
		vendor_master.registered_by = frappe.session.user

	def update_vendor_master_fields(self, vendor_master, mapped_row):
		"""Update fields for existing vendor master"""
		
		# Update only non-empty fields
		update_fields = {
			'office_email_secondary': 'office_email_secondary',
			'mobile_number': 'mobile_number',
			'country': 'country'
		}
		
		for csv_field, doc_field in update_fields.items():
			value = self.safe_get_value(mapped_row, csv_field)
			if value:
				vendor_master.set(doc_field, value)

		vtype_row = {
				"vendor_type": self.safe_get_value(mapped_row, 'vendor_type')
			}
		vendor_master.append("vendor_types", vtype_row)

	def apply_field_mapping(self, row, field_mapping):
		"""Apply field mapping to a data row"""
		mapped_row = {}
		
		for csv_field, system_field in field_mapping.items():
			if system_field and csv_field in row:
				# Use safe_get_value to handle float/int conversion
				value = self.safe_get_value(row, csv_field)
				if value:  # Only add non-empty values
					mapped_row[system_field] = value
		
		return mapped_row

	def validate_vendor_data(self, vendor_data):
		"""Validate each vendor record using current field mapping"""
		results = {
			"total_records": len(vendor_data),
			"valid_records": 0,
			"invalid_records": 0,
			"errors": [],
			"warnings": []
		}
		
		field_mapping = json.loads(self.field_mapping) if self.field_mapping else {}
		
		for idx, row in enumerate(vendor_data, 1):
			mapped_row = self.apply_field_mapping(row, field_mapping)
			
			# Validate required fields
			errors = []
			warnings = []
			
			# Check required fields with safe string conversion
			vendor_name = self.safe_get_value(mapped_row, 'vendor_name')
			if not vendor_name:
				errors.append(f"Row {idx}: Vendor name is required")
			
			vendor_code = self.safe_get_value(mapped_row, 'vendor_code')
			if not vendor_code:
				warnings.append(f"Row {idx}: Vendor code is missing")
			
			company_code = self.safe_get_value(mapped_row, 'company_code')
			if not company_code:
				warnings.append(f"Row {idx}: Company code is missing")
			
			# Validate email format
			email = self.safe_get_value(mapped_row, 'office_email_primary')
			if email and not validate_email_address(email):
				errors.append(f"Row {idx}: Invalid email format - {email}")
			
			# Validate GST number format (basic check)
			gst = self.safe_get_value(mapped_row, 'gst_no')
			if gst and not self.validate_gst_format(gst):
				warnings.append(f"Row {idx}: Invalid GST format - {gst}")
			
			# Record results
			if errors:
				results["invalid_records"] += 1
				results["errors"].extend(errors)
			else:
				results["valid_records"] += 1
			
			results["warnings"].extend(warnings)
		
		return results

	def validate_gst_format(self, gst):
		"""Basic GST format validation"""
		if not gst:
			return True  # GST is optional
		
		gst_str = str(gst).strip()
		# Basic GST format: 15 characters
		return len(gst_str) == 15 and gst_str.isalnum()

	def get_or_create_state(self, state_name):
		"""Get or create state master"""
		state_name_str = self.safe_get_value({'state': state_name}, 'state')
		if not state_name_str:
			return None
		
		state = frappe.db.exists("State Master", {"state_name": state_name_str})
		if not state:
			try:
				state_doc = frappe.new_doc("State Master")
				state_doc.state_name = state_name_str
				state_doc.state_code = state_name_str[:2].upper()
				state_doc.insert(ignore_permissions=True)
				return state_doc.name
			except:
				return None
		return state

	def get_or_create_city(self, city_name):
		"""Get or create city master"""
		city_name_str = self.safe_get_value({'city': city_name}, 'city')
		if not city_name_str:
			return None
		
		city = frappe.db.exists("City Master", {"city_name": city_name_str})
		if not city:
			try:
				city_doc = frappe.new_doc("City Master")
				city_doc.city_name = city_name_str
				city_doc.insert(ignore_permissions=True)
				return city_doc.name
			except:
				return None
		return city

	def get_or_create_pincode(self, pincode):
		"""Get or create pincode master"""
		pincode_str = self.safe_get_value({'pincode': pincode}, 'pincode')
		if not pincode_str:
			return None
		
		pincode_master = frappe.db.exists("Pincode Master", {"pincode": pincode_str})
		if not pincode_master:
			try:
				pincode_doc = frappe.new_doc("Pincode Master")
				pincode_doc.pincode = pincode_str
				pincode_doc.insert(ignore_permissions=True)
				return pincode_doc.name
			except:
				return None
		return pincode_master

	# HTML Generation Methods
	def generate_field_mapping_html(self, csv_headers):
		"""Generate HTML interface for field mapping"""
		
		target_fields = self.get_all_target_fields()
		auto_mapping = self.generate_auto_mapping(csv_headers)
		
		# Calculate mapping statistics
		mapped_count = sum(1 for v in auto_mapping.values() if v)
		unmapped_count = len(auto_mapping) - mapped_count
		mapping_percentage = (mapped_count / len(auto_mapping) * 100) if auto_mapping else 0
		
		html = f"""
		<div class="field-mapping-container">
			<div class="card">
				<div class="card-header bg-primary text-white">
					<h5 class="mb-0"><i class="fa fa-exchange-alt"></i> Field Mapping Configuration</h5>
					<small>Map your CSV columns to system fields. Auto-mapping has been applied based on header similarity.</small>
				</div>
				<div class="card-body">
					<!-- Mapping Statistics -->
					<div class="mapping-stats-section mb-4">
						<div class="row">
							<div class="col-md-8">
								<h6>Mapping Statistics</h6>
								<div class="progress mb-2" style="height: 25px;">
									<div class="progress-bar bg-success" role="progressbar" 
										style="width: {mapping_percentage}%" 
										aria-valuenow="{mapping_percentage}" 
										aria-valuemin="0" 
										aria-valuemax="100">
										{mapping_percentage:.1f}% Mapped
									</div>
								</div>
								<div class="mapping-summary">
									<span class="badge bg-success mapped-count">{mapped_count}</span> Mapped &nbsp;
									<span class="badge bg-warning unmapped-count">{unmapped_count}</span> Unmapped &nbsp;
									<span class="badge bg-info total-count">{len(auto_mapping)}</span> Total
								</div>
							</div>
							<div class="col-md-4">
								<div class="mapping-actions">
									<button type="button" class="btn btn-info btn-sm mb-1" onclick="apply_auto_mapping()">
										<i class="fa fa-magic"></i> Re-apply Auto Mapping
									</button><br>
									<button type="button" class="btn btn-warning btn-sm" onclick="clear_all_mapping()">
										<i class="fa fa-eraser"></i> Clear All Mapping
									</button>
								</div>
							</div>
						</div>
					</div>
					
					<div class="table-responsive">
						<table class="table table-bordered table-hover mapping-table">
							<thead class="table-dark">
								<tr>
									<th width="5%">#</th>
									<th width="35%">CSV Column Header</th>
									<th width="35%">Map to System Field</th>
									<th width="15%">Status</th>
									<th width="10%">Sample Data</th>
								</tr>
							</thead>
							<tbody>
		"""
		
		# Get sample data for preview
		sample_data = {}
		if self.vendor_data:
			vendor_data = json.loads(self.vendor_data)
			if vendor_data:
				sample_data = vendor_data[0]
		
		for idx, header in enumerate(csv_headers):
			auto_mapped_field = auto_mapping.get(header, '')
			status_class = 'success' if auto_mapped_field else 'warning'
			status_text = 'Mapped' if auto_mapped_field else 'Unmapped'
			
			# Get sample data for this column
			sample_value = sample_data.get(header, 'N/A')
			if pd.isna(sample_value):
				sample_value = 'N/A'
			else:
				sample_value = str(sample_value)[:20] + '...' if len(str(sample_value)) > 20 else str(sample_value)
			
			options_html = '<option value="">-- Select Field --</option>'
			for field_group, fields in target_fields.items():
				options_html += f'<optgroup label="{field_group}">'
				for field_key, field_label in fields.items():
					selected = 'selected' if field_key == auto_mapped_field else ''
					options_html += f'<option value="{field_key}" {selected}>{field_label}</option>'
				options_html += '</optgroup>'
			
			html += f"""
				<tr>
					<td>{idx + 1}</td>
					<td><strong>{header}</strong></td>
					<td>
						<select class="form-control field-mapping-select" 
								data-csv-field="{header}" 
								onchange="update_field_mapping(this)">
							{options_html}
						</select>
					</td>
					<td>
						<span class="badge bg-{status_class}">{status_text}</span>
					</td>
					<td>
						<small class="text-muted">{sample_value}</small>
					</td>
				</tr>
			"""
		
		html += """
							</tbody>
						</table>
					</div>
				</div>
			</div>
		</div>
		
		<script>
			function update_field_mapping(selectElement) {
				// Get current mapping
				let currentMapping = {};
				try {
					currentMapping = JSON.parse(cur_frm.doc.field_mapping || '{}');
				} catch(e) {
					currentMapping = {};
				}
				
				// Update mapping
				let csvField = selectElement.getAttribute('data-csv-field');
				let systemField = selectElement.value;
				
				if (systemField) {
					currentMapping[csvField] = systemField;
				} else {
					delete currentMapping[csvField];
				}
				
				// Save back to form
				cur_frm.set_value('field_mapping', JSON.stringify(currentMapping, null, 2));
				
				// Update statistics
				update_mapping_statistics();
			}
			
			function update_mapping_statistics() {
				let mapping = {};
				try {
					mapping = JSON.parse(cur_frm.doc.field_mapping || '{}');
				} catch(e) {
					mapping = {};
				}
				
				let total = Object.keys(mapping).length;
				let mapped = Object.values(mapping).filter(v => v).length;
				let unmapped = total - mapped;
				let percentage = total > 0 ? (mapped / total * 100) : 0;
				
				// Update display
				$('.mapped-count').text(mapped);
				$('.unmapped-count').text(unmapped);
				$('.total-count').text(total);
				$('.progress-bar').css('width', percentage + '%').text(percentage.toFixed(1) + '% Mapped');
				
				// Update row statuses
				$('.field-mapping-select').each(function() {
					let row = $(this).closest('tr');
					let badge = row.find('.badge');
					
					if ($(this).val()) {
						badge.removeClass('bg-warning').addClass('bg-success').text('Mapped');
					} else {
						badge.removeClass('bg-success').addClass('bg-warning').text('Unmapped');
					}
				});
			}
			
			function apply_auto_mapping() {
				frappe.call({
					method: 'vms.vendor_onboarding.doctype.existing_vendor_import.existing_vendor_import.get_auto_mapping',
					args: { docname: cur_frm.doc.name },
					callback: function(r) {
						if (r.message) {
							cur_frm.set_value('field_mapping', JSON.stringify(r.message, null, 2));
							cur_frm.save().then(() => {
								frappe.show_alert({
									message: __('Auto mapping applied successfully'),
									indicator: 'green'
								});
								location.reload();
							});
						}
					}
				});
			}
			
			function clear_all_mapping() {
				frappe.confirm('Are you sure you want to clear all field mappings?', function() {
					cur_frm.set_value('field_mapping', '{}');
					cur_frm.save().then(() => {
						location.reload();
					});
				});
			}
		</script>
		
		<style>
			.field-mapping-container .card {
				border-radius: 10px;
				box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
			}
			
			.mapping-table th {
				background-color: #343a40 !important;
				color: white;
				border-color: #454d55;
			}
			
			.mapping-table td {
				vertical-align: middle;
			}
			
			.field-mapping-select {
				border-radius: 6px;
				border: 2px solid #dee2e6;
				transition: border-color 0.3s ease;
			}
			
			.field-mapping-select:focus {
				border-color: #007bff;
				box-shadow: 0 0 0 0.2rem rgba(0, 123, 255, 0.25);
			}
			
			.mapping-summary .badge {
				margin-right: 10px;
				padding: 8px 12px;
				font-size: 0.9rem;
			}
			
			.mapping-actions .btn {
				width: 100%;
				margin-bottom: 5px;
			}
			
			.progress {
				border-radius: 10px;
			}
			
			.progress-bar {
				transition: width 0.6s ease;
				font-weight: bold;
			}
		</style>
		"""
		
		self.field_mapping_html = html

	def get_all_target_fields(self):
		"""Get all available target fields organized by DocType"""
		
		return {
			"Vendor Master": {
				"vendor_name": "Vendor Name",
				"office_email_primary": "Primary Email",
				"office_email_secondary": "Secondary Email", 
				"mobile_number": "Mobile Number",
				"country": "Country",
				"payee_in_document": "Payee in Document",
				"gr_based_inv_ver": "GR Based Invoice Verification",
				"service_based_inv_ver": "Service Based Invoice Verification",
				"check_double_invoice": "Check Double Invoice"
			},
			
			"Company Details": {
				"company_name": "Company Name",
				"gst": "GST Number",
				"company_pan_number": "PAN Number",
				"address_line_1": "Address Line 1",
				"address_line_2": "Address Line 2",
				"city": "City",
				"state": "State",
				"country": "Country",
				"pincode": "Pincode",
				"telephone_number": "Telephone Number",
				"nature_of_business": "Nature of Business",
				"type_of_business": "Type of Business",
				"corporate_identification_number": "CIN Number",
				"established_year": "Established Year"
			},
			
			"Company Vendor Code": {
				"company_code": "Company Code",
				"vendor_code": "Vendor Code", 
				"gst_no": "GST Number",
				"state": "State"
			},
			
			"Multiple Company Data": {
				"purchase_organization": "Purchase Organization",
				"account_group": "Account Group",
				"terms_of_payment": "Terms of Payment",
				"purchase_group": "Purchase Group",
				"order_currency": "Order Currency",
				"incoterms": "Incoterms",
				"reconciliation_account": "Reconciliation Account"
			},
			
			"Payment Details": {
				"bank_name": "Bank Name",
				"ifsc_code": "IFSC Code",
				"account_number": "Account Number",
				"name_of_account_holder": "Account Holder Name",
				"type_of_account": "Account Type"
			},
			
			"Additional Fields": {
				"vendor_gst_classification": "GST Classification",
				"nature_of_services": "Nature of Services",
				"vendor_type": "Vendor Type",
				"remarks": "Remarks",
				"contact_person": "Contact Person",
				"hod": "HOD",
				"enterprise_registration_no": "Enterprise Registration Number"
			}
		}

	def generate_mapping_statistics_html(self):
		"""Generate mapping statistics HTML"""
		if not self.field_mapping:
			return "<div class='alert alert-info'>Upload a file to see mapping statistics</div>"
		
		try:
			field_mapping = json.loads(self.field_mapping)
			
			# Calculate statistics
			total_fields = len(field_mapping)
			mapped_fields = sum(1 for v in field_mapping.values() if v)
			unmapped_fields = total_fields - mapped_fields
			mapping_percentage = (mapped_fields / total_fields * 100) if total_fields > 0 else 0
			
			html = f"""
			<div class="card">
				<div class="card-header bg-info text-white">
					<h6><i class="fa fa-chart-pie"></i> Mapping Statistics</h6>
				</div>
				<div class="card-body">
					<div class="progress mb-3" style="height: 25px;">
						<div class="progress-bar bg-success" style="width: {mapping_percentage}%">
							{mapping_percentage:.1f}% Mapped
						</div>
					</div>
					<div class="row text-center">
						<div class="col-md-4">
							<h4 class="text-primary">{total_fields}</h4>
							<small>Total Fields</small>
						</div>
						<div class="col-md-4">
							<h4 class="text-success">{mapped_fields}</h4>
							<small>Mapped</small>
						</div>
						<div class="col-md-4">
							<h4 class="text-warning">{unmapped_fields}</h4>
							<small>Unmapped</small>
						</div>
					</div>
				</div>
			</div>
			"""
			return html
			
		except Exception as e:
			return f"<div class='alert alert-danger'>Error: {str(e)}</div>"

	def generate_display_html(self, vendor_data, validation_results):
		"""Generate all display HTML sections"""
		
		# Generate success/fail rate HTML
		success_rate = (validation_results['valid_records'] / validation_results['total_records'] * 100) if validation_results['total_records'] > 0 else 0
		fail_rate = 100 - success_rate
		
		success_html = f"""
		<div class="import-summary-container">
			<div class="row">
				<div class="col-md-8">
					<div class="row">
						<div class="col-md-3">
							<div class="summary-stat">
								<div class="stat-circle bg-primary">
									<div class="stat-number">{validation_results['total_records']}</div>
								</div>
								<div class="stat-label">Total Records</div>
							</div>
						</div>
						<div class="col-md-3">
							<div class="summary-stat">
								<div class="stat-circle bg-success">
									<div class="stat-number">{validation_results['valid_records']}</div>
								</div>
								<div class="stat-label">Valid Records</div>
							</div>
						</div>
						<div class="col-md-3">
							<div class="summary-stat">
								<div class="stat-circle bg-danger">
									<div class="stat-number">{validation_results['invalid_records']}</div>
								</div>
								<div class="stat-label">Invalid Records</div>
							</div>
						</div>
						<div class="col-md-3">
							<div class="summary-stat">
								<div class="stat-circle bg-info">
									<div class="stat-number">{len(json.loads(self.field_mapping) if self.field_mapping else {})}</div>
								</div>
								<div class="stat-label">CSV Columns</div>
							</div>
						</div>
					</div>
				</div>
				<div class="col-md-4">
					<div class="progress-section">
						<h6>Success Rate</h6>
						<div class="progress mb-2" style="height: 20px;">
							<div class="progress-bar bg-success" style="width: {success_rate}%">{success_rate:.1f}%</div>
						</div>
						<div class="progress" style="height: 20px;">
							<div class="progress-bar bg-danger" style="width: {fail_rate}%">{fail_rate:.1f}% Failed</div>
						</div>
					</div>
				</div>
			</div>
			
			{self.generate_validation_charts(validation_results)}
			{self.generate_errors_html(validation_results)}
		</div>
		"""
		
		# Enhanced Vendor Data HTML with schema
		vendor_html = self.generate_vendor_data_schema_html(vendor_data, validation_results)
		
		# Add CSS
		css = """
		<style>
			.summary-stat {
				margin-bottom: 20px;
			}
			
			.stat-circle {
				width: 80px;
				height: 80px;
				border-radius: 50%;
				display: flex;
				align-items: center;
				justify-content: center;
				margin: 0 auto 10px auto;
				color: white;
			}
			
			.stat-number {
				font-size: 1.5rem;
				font-weight: bold;
			}
			
			.stat-label {
				font-weight: 500;
				color: #495057;
			}
			
			.progress-section {
				background: #f8f9fa;
				padding: 15px;
				border-radius: 8px;
				border: 1px solid #dee2e6;
			}
			
			.download-actions .btn {
				width: 100%;
			}
			
			.validation-charts {
				background: #fff;
				padding: 20px;
				border-radius: 8px;
				border: 1px solid #dee2e6;
				margin-top: 20px;
			}
		</style>
		"""
		
		self.success_fail_rate_html = success_html + css
		self.vendor_html = vendor_html

	def generate_validation_charts(self, validation_results):
		"""Generate validation charts section"""
		if not validation_results:
			return ""
		
		errors_count = len(validation_results.get('errors', []))
		warnings_count = len(validation_results.get('warnings', []))
		
		html = f"""
		<div class="validation-charts">
			<div class="row">
				<div class="col-md-6">
					<h6>Validation Results</h6>
					<canvas id="validationResultsChart" width="300" height="200"></canvas>
				</div>
				<div class="col-md-6">
					<h6>Issue Breakdown</h6>
					<div class="issue-stats">
						<div class="issue-item">
							<div class="issue-icon bg-danger">
								<i class="fa fa-times"></i>
							</div>
							<div class="issue-details">
								<div class="issue-count">{errors_count}</div>
								<div class="issue-label">Errors</div>
							</div>
						</div>
						<div class="issue-item">
							<div class="issue-icon bg-warning">
								<i class="fa fa-exclamation-triangle"></i>
							</div>
							<div class="issue-details">
								<div class="issue-count">{warnings_count}</div>
								<div class="issue-label">Warnings</div>
							</div>
						</div>
					</div>
				</div>
			</div>
		</div>
		
		<script>
		setTimeout(function() {{
			// Create validation results chart
			const ctx = document.getElementById('validationResultsChart');
			if (ctx) {{
				new Chart(ctx, {{
					type: 'doughnut',
					data: {{
						labels: ['Valid Records', 'Invalid Records'],
						datasets: [{{
							data: [{validation_results['valid_records']}, {validation_results['invalid_records']}],
							backgroundColor: ['#28a745', '#dc3545'],
							borderWidth: 2,
							borderColor: '#fff'
						}}]
					}},
					options: {{
						responsive: true,
						plugins: {{
							legend: {{
								position: 'bottom'
							}}
						}}
					}}
				}});
			}}
		}}, 1000);
		</script>
		
		<style>
			.issue-stats {{
				display: flex;
				gap: 20px;
			}}
			
			.issue-item {{
				display: flex;
				align-items: center;
				gap: 10px;
				flex: 1;
			}}
			
			.issue-icon {{
				width: 40px;
				height: 40px;
				border-radius: 50%;
				display: flex;
				align-items: center;
				justify-content: center;
				color: white;
			}}
			
			.issue-count {{
				font-size: 1.5rem;
				font-weight: bold;
			}}
			
			.issue-label {{
				color: #6c757d;
				font-size: 0.9rem;
			}}
		</style>
		"""
		
		return html

	def generate_errors_html(self, validation_results):
		"""Generate errors and warnings HTML"""
		errors = validation_results.get('errors', [])
		warnings = validation_results.get('warnings', [])
		
		if not errors and not warnings:
			return '<div class="alert alert-success mt-3"><i class="fa fa-check-circle"></i> All records passed validation!</div>'
		
		html = '<div class="validation-issues mt-3">'
		
		if errors:
			html += f"""
			<div class="card border-danger mb-3">
				<div class="card-header bg-danger text-white">
					<h6 class="mb-0"><i class="fa fa-times-circle"></i> Errors ({len(errors)})</h6>
				</div>
				<div class="card-body">
					<div class="error-list" style="max-height: 200px; overflow-y: auto;">
			"""
			for error in errors[:20]:  # Show first 20 errors
				html += f'<div class="alert alert-danger py-2 mb-1"><small>{error}</small></div>'
			
			if len(errors) > 20:
				html += f'<div class="alert alert-info py-2"><small>... and {len(errors) - 20} more errors</small></div>'
			
			html += '</div></div></div>'
		
		if warnings:
			html += f"""
			<div class="card border-warning mb-3">
				<div class="card-header bg-warning text-dark">
					<h6 class="mb-0"><i class="fa fa-exclamation-triangle"></i> Warnings ({len(warnings)})</h6>
				</div>
				<div class="card-body">
					<div class="warning-list" style="max-height: 200px; overflow-y: auto;">
			"""
			for warning in warnings[:20]:  # Show first 20 warnings
				html += f'<div class="alert alert-warning py-2 mb-1"><small>{warning}</small></div>'
			
			if len(warnings) > 20:
				html += f'<div class="alert alert-info py-2"><small>... and {len(warnings) - 20} more warnings</small></div>'
			
			html += '</div></div></div>'
		
		html += '</div>'
		return html

	def generate_vendor_data_schema_html(self, vendor_data, validation_results):
		"""Generate enhanced vendor data HTML with schema and graph format"""
		
		if not vendor_data:
			return "<div class='alert alert-info'>No vendor data available</div>"
		
		field_mapping = json.loads(self.field_mapping) if self.field_mapping else {}
		
		total_records = len(vendor_data)
		valid_records = validation_results.get('valid_records', 0)
		invalid_records = validation_results.get('invalid_records', 0)
		total_columns = len(field_mapping)
		
		html = f"""
		<div class="vendor-data-container">
			<!-- Data Overview -->
			<div class="card mb-3">
				<div class="card-header bg-primary text-white">
					<h5 class="mb-0"><i class="fa fa-database"></i> Vendor Data Schema & Preview</h5>
				</div>
				<div class="card-body">
					<div class="row">
						<div class="col-md-8">
							<div class="data-stats">
								<div class="row">
									<div class="col-md-3">
										<div class="stat-box bg-primary">
											<div class="stat-number">{total_records}</div>
											<div class="stat-label">Total Records</div>
										</div>
									</div>
									<div class="col-md-3">
										<div class="stat-box bg-success">
											<div class="stat-number">{valid_records}</div>
											<div class="stat-label">Valid Records</div>
										</div>
									</div>
									<div class="col-md-3">
										<div class="stat-box bg-danger">
											<div class="stat-number">{invalid_records}</div>
											<div class="stat-label">Invalid Records</div>
										</div>
									</div>
									<div class="col-md-3">
										<div class="stat-box bg-info">
											<div class="stat-number">{total_columns}</div>
											<div class="stat-label">CSV Columns</div>
										</div>
									</div>
								</div>
							</div>
						</div>
						<div class="col-md-4">
							<div class="validation-chart">
								<canvas id="validationChart" width="200" height="200"></canvas>
							</div>
						</div>
					</div>
				</div>
			</div>
			
			<!-- Data Schema Visualization -->
			<div class="card mb-3">
				<div class="card-header bg-info text-white">
					<h6 class="mb-0"><i class="fa fa-sitemap"></i> Data Flow Schema</h6>
				</div>
				<div class="card-body">
					<div class="schema-diagram">
						{self.generate_schema_diagram()}
					</div>
				</div>
			</div>
			
			<!-- Detailed Data Table -->
			<div class="card">
				<div class="card-header bg-dark text-white">
					<h6 class="mb-0"><i class="fa fa-table"></i> Vendor Records Preview</h6>
				</div>
				<div class="card-body">
					<div class="table-responsive">
						<table class="table table-striped table-bordered vendor-data-table">
							<thead class="table-dark">
								<tr>
									<th width="5%">ID</th>
									<th width="15%">Vendor Name</th>
									<th width="10%">Code</th>
									<th width="10%">Company</th>
									<th width="10%">State</th>
									<th width="15%">GST</th>
									<th width="15%">Email</th>
									<th width="10%">Phone</th>
									<th width="10%">Status</th>
								</tr>
							</thead>
							<tbody>
		"""
		
		# Display actual vendor data
		for idx, row in enumerate(vendor_data[:10], 1):  # Show first 10 records
			mapped_row = self.apply_field_mapping(row, field_mapping)
			
			# Determine status
			is_valid = idx <= valid_records
			status_class = "success" if is_valid else "danger"
			status_text = "Valid" if is_valid else "Invalid"
			
			# Get values with fallbacks using safe string conversion
			vendor_name = self.safe_get_value(mapped_row, 'vendor_name') or 'N/A'
			vendor_code = self.safe_get_value(mapped_row, 'vendor_code') or 'N/A'
			company_code = self.safe_get_value(mapped_row, 'company_code') or 'N/A'
			state = self.safe_get_value(mapped_row, 'state') or 'N/A'
			gst = self.safe_get_value(mapped_row, 'gst') or self.safe_get_value(mapped_row, 'gst_no') or 'N/A'
			email = self.safe_get_value(mapped_row, 'office_email_primary') or 'N/A'
			phone = self.safe_get_value(mapped_row, 'mobile_number') or 'N/A'
			
			# Truncate long values
			vendor_name = (vendor_name[:20] + '...') if len(vendor_name) > 20 else vendor_name
			email = (email[:20] + '...') if len(email) > 20 else email
			
			html += f"""
				<tr class="vendor-row">
					<td><span class="badge bg-secondary">{idx}</span></td>
					<td><strong>{vendor_name}</strong></td>
					<td><span class="vendor-code">{vendor_code}</span></td>
					<td><span class="company-code">{company_code}</span></td>
					<td><span class="state-name">{state}</span></td>
					<td><span class="gst-number">{gst}</span></td>
					<td><span class="email-address">{email}</span></td>
					<td><span class="phone-number">{phone}</span></td>
					<td><span class="badge bg-{status_class}">{status_text}</span></td>
				</tr>
			"""
		
		# Show more records indicator
		if len(vendor_data) > 10:
			html += f"""
				<tr>
					<td colspan="9" class="text-center text-muted">
						<i class="fa fa-ellipsis-h"></i> ... and {len(vendor_data) - 10} more records
					</td>
				</tr>
			"""
		
		html += """
							</tbody>
						</table>
					</div>
				</div>
			</div>
		</div>
		
		<style>
			.vendor-data-container .stat-box {
				text-align: center;
				padding: 15px;
				border-radius: 8px;
				color: white;
				margin-bottom: 10px;
			}
			
			.vendor-data-container .stat-number {
				font-size: 1.8rem;
				font-weight: bold;
				display: block;
			}
			
			.vendor-data-container .stat-label {
				font-size: 0.9rem;
				opacity: 0.9;
			}
			
			.vendor-row:hover {
				background-color: #f8f9fa;
			}
			
			.vendor-code, .company-code {
				background: #e3f2fd;
				padding: 2px 6px;
				border-radius: 3px;
				font-family: monospace;
				font-size: 0.85rem;
			}
			
			.state-name {
				background: #e3f2fd;
				padding: 2px 6px;
				border-radius: 3px;
				font-size: 0.8rem;
			}
			
			.gst-number, .email-address, .phone-number {
				font-family: monospace;
			}
		</style>
		"""
		
		return html

	def generate_schema_diagram(self):
		"""Generate a visual schema diagram"""
		return """
		<div class="schema-flow">
			<div class="schema-step">
				<div class="step-box csv-box">
					<i class="fa fa-file-csv fa-2x"></i>
					<div class="step-title">CSV Data</div>
					<div class="step-desc">Raw vendor data</div>
				</div>
			</div>
			<div class="schema-arrow">→</div>
			<div class="schema-step">
				<div class="step-box mapping-box">
					<i class="fa fa-exchange-alt fa-2x"></i>
					<div class="step-title">Field Mapping</div>
					<div class="step-desc">CSV to System fields</div>
				</div>
			</div>
			<div class="schema-arrow">→</div>
			<div class="schema-step">
				<div class="step-box validation-box">
					<i class="fa fa-check-circle fa-2x"></i>
					<div class="step-title">Validation</div>
					<div class="step-desc">Data quality checks</div>
				</div>
			</div>
			<div class="schema-arrow">→</div>
			<div class="schema-step">
				<div class="step-box processing-box">
					<i class="fa fa-cogs fa-2x"></i>
					<div class="step-title">Processing</div>
					<div class="step-desc">Create vendor records</div>
				</div>
			</div>
		</div>
		
		<div class="data-flow-details mt-3">
			<div class="row">
				<div class="col-md-3">
					<div class="flow-detail">
						<h6>Vendor Master</h6>
						<small>Basic vendor info</small>
					</div>
				</div>
				<div class="col-md-3">
					<div class="flow-detail">
						<h6>Company Data</h6>
						<small>Multi-company relationships</small>
					</div>
				</div>
				<div class="col-md-3">
					<div class="flow-detail">
						<h6>Vendor Codes</h6>
						<small>SAP codes by state</small>
					</div>
				</div>
				<div class="col-md-3">
					<div class="flow-detail">
						<h6>Company Details</h6>
						<small>Address & business info</small>
					</div>
				</div>
			</div>
		</div>
		
		<style>
			.schema-flow {
				display: flex;
				align-items: center;
				justify-content: center;
				flex-wrap: wrap;
				gap: 20px;
			}
			
			.step-box {
				padding: 20px;
				border-radius: 10px;
				text-align: center;
				color: white;
				min-width: 120px;
			}
			
			.csv-box { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); }
			.mapping-box { background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); }
			.validation-box { background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%); }
			.processing-box { background: linear-gradient(135deg, #43e97b 0%, #38f9d7 100%); }
			
			.step-title {
				font-weight: bold;
				margin: 10px 0 5px 0;
			}
			
			.step-desc {
				font-size: 0.8rem;
				opacity: 0.9;
			}
			
			.schema-arrow {
				font-size: 1.5rem;
				color: #6c757d;
				font-weight: bold;
			}
			
			.flow-detail {
				background: #f8f9fa;
				padding: 10px;
				border-radius: 6px;
				border: 1px solid #dee2e6;
				text-align: center;
			}
			
			.flow-detail h6 {
				color: #495057;
				margin-bottom: 5px;
			}
			
			@media (max-width: 768px) {
				.schema-flow {
					flex-direction: column;
				}
				.schema-arrow {
					transform: rotate(90deg);
				}
			}
		</style>
		"""


# API Methods
@frappe.whitelist()
def process_existing_vendors(docname):
	"""API method to process vendor import"""
	doc = frappe.get_doc("Existing Vendor Import", docname)
	return doc.process_vendors()


@frappe.whitelist()
def get_auto_mapping(docname):
	"""Get automatic field mapping for CSV headers"""
	doc = frappe.get_doc("Existing Vendor Import", docname)
	
	if not doc.original_headers:
		return {}
	
	headers = json.loads(doc.original_headers)
	return doc.generate_auto_mapping(headers)


@frappe.whitelist()
def download_processed_data(docname, data_type="all"):
	"""Download processed vendor data as Excel"""
	doc = frappe.get_doc("Existing Vendor Import", docname)
	
	if not doc.vendor_data:
		frappe.throw("No vendor data found")
	
	vendor_data = json.loads(doc.vendor_data)
	field_mapping = json.loads(doc.field_mapping) if doc.field_mapping else {}
	validation_results = json.loads(doc.success_fail_rate) if doc.success_fail_rate else {}
	
	# Apply field mapping to all data
	processed_data = []
	for idx, row in enumerate(vendor_data):
		mapped_row = doc.apply_field_mapping(row, field_mapping)
		mapped_row['_row_number'] = idx + 1
		mapped_row['_status'] = 'Valid' if idx < validation_results.get('valid_records', 0) else 'Invalid'
		processed_data.append(mapped_row)
	
	# Filter data based on type
	if data_type == "valid":
		processed_data = [row for row in processed_data if row['_status'] == 'Valid']
		filename = f"valid_vendors_{doc.name}.xlsx"
	elif data_type == "invalid":
		processed_data = [row for row in processed_data if row['_status'] == 'Invalid']
		filename = f"invalid_vendors_{doc.name}.xlsx"
	else:
		filename = f"all_vendors_{doc.name}.xlsx"
	
	# Create Excel file
	output = BytesIO()
	with pd.ExcelWriter(output, engine='openpyxl') as writer:
		# Main data sheet
		if processed_data:
			df = pd.DataFrame(processed_data)
			df.to_excel(writer, sheet_name='Vendor Data', index=False)
		
		# Field mapping sheet
		mapping_df = pd.DataFrame([
			{"CSV Header": k, "System Field": v or "Not Mapped"} 
			for k, v in field_mapping.items()
		])
		mapping_df.to_excel(writer, sheet_name='Field Mapping', index=False)
		
		# Validation results
		if validation_results.get('errors'):
			errors_df = pd.DataFrame([{"Error": error} for error in validation_results['errors']])
			errors_df.to_excel(writer, sheet_name='Errors', index=False)
		
		# Warnings sheet
		if validation_results.get('warnings'):
			warnings_df = pd.DataFrame([{"Warning": warning} for warning in validation_results['warnings']])
			warnings_df.to_excel(writer, sheet_name='Warnings', index=False)

	output.seek(0)

	# Save file
	from frappe.utils.file_manager import save_file
	file_doc = save_file(filename, output.read(), doc.doctype, doc.name, is_private=0)

	return {
		"file_url": file_doc.file_url,
		"file_name": filename
	}


@frappe.whitelist()
def download_field_mapping_template(docname):
	"""Download field mapping template"""
	try:
		doc = frappe.get_doc("Existing Vendor Import", docname)
		
		# Create Excel file with field mapping template
		output = io.BytesIO()
		workbook = openpyxl.Workbook()
		worksheet = workbook.active
		worksheet.title = "Field Mapping Template"
		
		# Headers
		headers = ["CSV Column", "System Field", "Description", "Required"]
		for col_num, header in enumerate(headers, 1):
			cell = worksheet.cell(row=1, column=col_num)
			cell.value = header
			cell.font = openpyxl.styles.Font(bold=True)
			cell.fill = openpyxl.styles.PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
		
		# Get all available target fields
		target_fields = doc.get_all_target_fields()
		
		row_num = 2
		for doctype_name, fields in target_fields.items():
			for field_name, field_label in fields.items():
				worksheet.cell(row=row_num, column=1, value="")  # CSV Column - to be filled by user
				worksheet.cell(row=row_num, column=2, value=field_name)
				worksheet.cell(row=row_num, column=3, value=f"{doctype_name}: {field_label}")
				worksheet.cell(row=row_num, column=4, value="No")  # Most fields are optional
				row_num += 1
		
		# Adjust column widths
		for column in worksheet.columns:
			max_length = 0
			column_letter = column[0].column_letter
			for cell in column:
				try:
					if len(str(cell.value)) > max_length:
						max_length = len(str(cell.value))
				except:
					pass
			adjusted_width = min(max_length + 2, 50)
			worksheet.column_dimensions[column_letter].width = adjusted_width
		
		workbook.save(output)
		output.seek(0)
		
		filename = f"field_mapping_template_{frappe.utils.now_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx"
		
		# Fix the attachment issue by providing proper attached_to_name
		file_doc = save_file(
			filename, 
			output.read(), 
			"Existing Vendor Import", 
			docname,  # Use the document name instead of None
			is_private=0
		)
		
		return {
			"file_url": file_doc.file_url,
			"file_name": filename
		}
		
	except Exception as e:
		frappe.log_error(f"Error downloading field mapping template: {str(e)}")
		frappe.throw(f"Error creating template: {str(e)}")

@frappe.whitelist()
def get_vendor_import_preview(docname):
	"""Get preview of vendor data with mapping applied"""
	doc = frappe.get_doc("Existing Vendor Import", docname)
	
	if not doc.vendor_data or not doc.field_mapping:
		return {"error": "Vendor data and field mapping required"}
	
	vendor_data = json.loads(doc.vendor_data)
	field_mapping = json.loads(doc.field_mapping)
	
	# Apply mapping to first 5 records for preview
	preview_data = []
	for idx, row in enumerate(vendor_data[:5]):
		mapped_row = doc.apply_field_mapping(row, field_mapping)
		mapped_row['_row_number'] = idx + 1
		preview_data.append(mapped_row)
	
	return {
		"preview_data": preview_data,
		"total_records": len(vendor_data),
		"mapped_fields": sum(1 for v in field_mapping.values() if v)
	}


# Add this method to the ExistingVendorImport class in existing_vendor_import.py

def check_for_duplicates(self, vendor_data):
	"""Check for duplicate vendors in the dataset using VendorImportUtils"""
	from .existing_vendor_import_utils import VendorImportUtils
	
	try:
		# Use the utility class method for checking duplicates
		duplicates = VendorImportUtils.get_duplicate_vendors(vendor_data)
		
		return {
			'duplicates': duplicates,
			'duplicate_count': len(duplicates),
			'has_duplicates': len(duplicates) > 0
		}
		
	except Exception as e:
		frappe.log_error(f"Error checking for duplicates: {str(e)}")
		return {
			'duplicates': [],
			'duplicate_count': 0,
			'has_duplicates': False,
			'error': str(e)
		}

# Also update the validate_import_data method around line 1894
@frappe.whitelist()
def validate_import_data(docname):
	"""Validate import data before processing"""
	try:
		doc = frappe.get_doc("Existing Vendor Import", docname)
		
		if not doc.vendor_data:
			return {"error": "No vendor data found"}
		
		vendor_data = json.loads(doc.vendor_data)
		field_mapping = json.loads(doc.field_mapping) if doc.field_mapping else {}
		
		# Basic validation results structure
		results = {
			"total_records": len(vendor_data),
			"valid_records": 0,
			"invalid_records": 0,
			"errors": [],
			"warnings": [],
			"duplicates": []
		}
		
		# Check for duplicates using the class method
		duplicate_results = doc.check_for_duplicates(vendor_data)
		if duplicate_results.get('duplicates'):
			results['duplicates'] = duplicate_results['duplicates']
		
		# Validate each record
		from .existing_vendor_import_utils import VendorImportUtils
		
		for idx, row in enumerate(vendor_data, 1):
			try:
				# Apply field mapping
				mapped_row = doc.apply_field_mapping(row, field_mapping)
				
				# Validate the mapped row
				validation_result = VendorImportUtils.validate_row_data(mapped_row, idx)
				
				if validation_result['is_valid']:
					results['valid_records'] += 1
				else:
					results['invalid_records'] += 1
				
				# Collect errors and warnings
				results['errors'].extend(validation_result.get('errors', []))
				results['warnings'].extend(validation_result.get('warnings', []))
				
			except Exception as e:
				results['errors'].append(f"Row {idx}: Validation error - {str(e)}")
				results['invalid_records'] += 1
		
		return results
		
	except Exception as e:
		frappe.log_error(f"Error validating import data: {str(e)}")
		return {"error": f"Validation failed: {str(e)}"}


@frappe.whitelist()
def get_import_summary(docname):
	"""Get comprehensive import summary"""
	doc = frappe.get_doc("Existing Vendor Import", docname)
	
	if not doc.vendor_data:
		return {"error": "No vendor data found"}
	
	vendor_data = json.loads(doc.vendor_data)
	field_mapping = json.loads(doc.field_mapping) if doc.field_mapping else {}
	
	# Calculate statistics
	summary = {
		"total_records": len(vendor_data),
		"mapped_fields": sum(1 for v in field_mapping.values() if v),
		"unmapped_fields": sum(1 for v in field_mapping.values() if not v),
		"field_mapping_percentage": 0,
		"companies": set(),
		"states": set(),
		"vendor_types": set()
	}
	
	if len(field_mapping) > 0:
		summary["field_mapping_percentage"] = (summary["mapped_fields"] / len(field_mapping)) * 100
	
	# Analyze data
	for row in vendor_data:
		mapped_row = doc.apply_field_mapping(row, field_mapping)
		
		if mapped_row.get('company_code'):
			summary["companies"].add(mapped_row['company_code'])
		if mapped_row.get('state'):
			summary["states"].add(mapped_row['state'])
		if mapped_row.get('vendor_type'):
			summary["vendor_types"].add(mapped_row['vendor_type'])
	
	# Convert sets to lists for JSON serialization
	summary["companies"] = list(summary["companies"])
	summary["states"] = list(summary["states"])
	summary["vendor_types"] = list(summary["vendor_types"])
	
	return summary



# Enhanced error handling methods to add to existing_vendor_import.py

def safe_get_value(self, data_dict, key, default_value=""):
	"""Safely get value from dictionary with proper string handling"""
	try:
		value = data_dict.get(key, default_value)
		if value is None or pd.isna(value):
			return default_value
		return str(value).strip() if value else default_value
	except Exception as e:
		frappe.log_error(f"Error getting value for key {key}: {str(e)}")
		return default_value

def validate_file_upload(self):
	"""Validate the uploaded CSV/Excel file"""
	if not self.csv_xl:
		frappe.throw("Please upload a CSV or Excel file")
	
	try:
		# Get file extension
		file_url = self.csv_xl
		file_extension = file_url.split('.')[-1].lower()
		
		if file_extension not in ['csv', 'xlsx', 'xls']:
			frappe.throw("Please upload a valid CSV or Excel file")
		
		# Check file size (optional - adjust as needed)
		try:
			file_doc = frappe.get_doc("File", {"file_url": file_url})
			if file_doc and hasattr(file_doc, 'file_size'):
				# Check if file is too large (e.g., > 10MB)
				if file_doc.file_size > 10 * 1024 * 1024:
					frappe.throw("File size too large. Please upload a file smaller than 10MB")
		except:
			pass  # File size check is optional
		
		return True
		
	except Exception as e:
		frappe.log_error(f"File validation error: {str(e)}")
		frappe.throw(f"File validation failed: {str(e)}")

def safe_json_loads(self, json_string, default_value=None):
	"""Safely parse JSON string with error handling"""
	if not json_string:
		return default_value or {}
	
	try:
		return json.loads(json_string)
	except (json.JSONDecodeError, TypeError, ValueError) as e:
		frappe.log_error(f"JSON parsing error: {str(e)}")
		return default_value or {}

# Override the validate method to add comprehensive validation
def validate(self):
	"""Enhanced validation method"""
	try:
		# Validate file upload
		if self.csv_xl:
			self.validate_file_upload()
		
		# Parse CSV data if file is uploaded but no vendor data exists
		if self.csv_xl and not self.vendor_data:
			try:
				self.parse_csv_data()
			except Exception as e:
				frappe.log_error(f"CSV parsing error: {str(e)}")
				frappe.msgprint(f"Error parsing CSV file: {str(e)}")
		
		# Generate field mapping HTML if we have data
		if self.vendor_data and not self.field_mapping_html:
			try:
				self.generate_field_mapping_html()
			except Exception as e:
				frappe.log_error(f"Field mapping HTML generation error: {str(e)}")
				
		# Validate and generate success/fail rate if we have mapping
		if self.vendor_data and self.field_mapping:
			try:
				self.validate_and_generate_reports()
			except Exception as e:
				frappe.log_error(f"Validation report generation error: {str(e)}")
		
	except Exception as e:
		frappe.log_error(f"Document validation error: {str(e)}")
		frappe.throw(f"Validation failed: {str(e)}")

# Add method to handle validation and reporting
def validate_and_generate_reports(self):
	"""Validate data and generate reports"""
	try:
		vendor_data = self.safe_json_loads(self.vendor_data, [])
		field_mapping = self.safe_json_loads(self.field_mapping, {})
		
		if not vendor_data:
			return
		
		# Run validation
		from .existing_vendor_import_utils import VendorImportUtils
		
		validation_results = {
			"total_records": len(vendor_data),
			"valid_records": 0,
			"invalid_records": 0,
			"errors": [],
			"warnings": []
		}
		
		for idx, row in enumerate(vendor_data, 1):
			try:
				mapped_row = self.apply_field_mapping(row, field_mapping)
				row_validation = VendorImportUtils.validate_row_data(mapped_row, idx)
				
				if row_validation.get('is_valid', False):
					validation_results['valid_records'] += 1
				else:
					validation_results['invalid_records'] += 1
				
				validation_results['errors'].extend(row_validation.get('errors', []))
				validation_results['warnings'].extend(row_validation.get('warnings', []))
				
			except Exception as e:
				validation_results['errors'].append(f"Row {idx}: {str(e)}")
				validation_results['invalid_records'] += 1
		
		# Generate HTML reports
		self.success_fail_rate = json.dumps(validation_results, indent=2, default=str)
		self.success_fail_rate_html = self.generate_validation_results_html(validation_results)
		self.vendor_html = self.generate_vendor_data_schema_html(vendor_data, validation_results)
		
	except Exception as e:
		frappe.log_error(f"Validation and reporting error: {str(e)}")
		# Don't throw here, just log the error
		pass